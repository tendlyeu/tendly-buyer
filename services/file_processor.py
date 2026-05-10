"""File upload handler and text extraction service for procurement documents."""

import logging
import os
import uuid
from pathlib import Path

logger = logging.getLogger(__name__)

UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")


# Stable error codes used by route handlers to look up translated messages.
# Keep these in sync with `procurements.upload_error_*` translation keys.
class UploadError(ValueError):
    def __init__(self, code: str, message: str = ""):
        self.code = code
        super().__init__(message or code)


class FileProcessor:
    ALLOWED_EXTENSIONS = {'.pdf', '.docx', '.doc', '.xlsx', '.xls', '.txt', '.csv'}
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

    MIME_TYPES = {
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.doc': 'application/msword',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.txt': 'text/plain',
        '.csv': 'text/csv',
    }

    async def process_upload(self, file, plan_id: str) -> dict:
        """Process an uploaded file: save to disk + extract text.

        Args:
            file: Starlette UploadFile object from request.form()
            plan_id: Procurement plan ID for directory organization

        Returns:
            dict with: file_path, file_name, file_size, mime_type, content_text
        """
        filename = file.filename or "untitled"
        ext = os.path.splitext(filename)[1].lower()

        # Validate extension
        if ext not in self.ALLOWED_EXTENSIONS:
            raise UploadError(
                "bad_extension",
                f"File type '{ext}' is not allowed. Supported: {', '.join(sorted(self.ALLOWED_EXTENSIONS))}"
            )

        # Read file content into memory to check size
        content = await file.read()
        file_size = len(content)

        if file_size > self.MAX_FILE_SIZE:
            raise UploadError(
                "too_large",
                f"File size ({file_size / (1024*1024):.1f} MB) exceeds maximum allowed ({self.MAX_FILE_SIZE / (1024*1024):.0f} MB)."
            )

        if file_size == 0:
            raise UploadError("empty", "Uploaded file is empty.")

        # Create directory for this plan
        plan_dir = os.path.join(UPLOADS_DIR, plan_id)
        os.makedirs(plan_dir, exist_ok=True)

        # Generate unique filename to avoid collisions
        unique_name = f"{uuid.uuid4().hex[:8]}_{filename}"
        file_path = os.path.join(plan_dir, unique_name)

        # Write to disk
        with open(file_path, "wb") as f:
            f.write(content)

        # Determine MIME type
        mime_type = self.MIME_TYPES.get(ext, "application/octet-stream")

        # Extract text. Failure here is non-fatal — the file is already on
        # disk and the user can re-upload or the AI review can fall back to
        # a "no extracted text" branch.
        content_text = ""
        try:
            content_text = self._extract_text(file_path, ext)
        except Exception as e:
            logger.warning("Text extraction failed for %s: %s", filename, e)

        return {
            "file_path": file_path,
            "file_name": filename,
            "file_size": file_size,
            "mime_type": mime_type,
            "content_text": content_text,
        }

    def _extract_text(self, file_path: str, ext: str) -> str:
        """Route to the appropriate text extractor based on file extension."""
        if ext == '.pdf':
            return self.extract_text_from_pdf(file_path)
        elif ext in ('.docx', '.doc'):
            return self.extract_text_from_docx(file_path)
        elif ext in ('.xlsx', '.xls'):
            return self.extract_text_from_xlsx(file_path)
        elif ext in ('.txt', '.csv'):
            return self.extract_text_from_txt(file_path)
        return ""

    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF using pdfplumber."""
        import pdfplumber

        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n\n".join(text_parts)

    def extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX using python-docx."""
        from docx import Document

        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def extract_text_from_xlsx(self, file_path: str) -> str:
        """Extract text from XLSX using openpyxl."""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells).strip()
                if line and line != " | ".join([""] * len(cells)):
                    text_parts.append(line)
        wb.close()
        return "\n".join(text_parts)

    def extract_text_from_txt(self, file_path: str) -> str:
        """Read plain text or CSV file."""
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()

    @staticmethod
    def delete_file(file_path: str) -> bool:
        """Delete a file from disk. Returns True if deleted, False if not found."""
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                return True
        except OSError as e:
            logger.warning("Could not delete file %s: %s", file_path, e)
        return False
